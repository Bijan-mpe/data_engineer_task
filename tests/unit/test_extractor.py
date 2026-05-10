"""
Unit tests for src.pipeline.extractor.

Tests are split into three groups:
  1. Helper function tests — pure-Python, no file I/O.
  2. Integration-style tests — parse the real .xlsm fixtures in data/.
  3. Error-handling tests — bad inputs raise ExtractionError cleanly.

No database or FastAPI dependencies are exercised here.
"""

from __future__ import annotations

from pathlib import Path

import pytest

from src.models.schemas import ExtractedFile
from src.pipeline.extractor import (
    ExtractionError,
    _build_label_map,
    _extract_year_specs,
    _is_metric_row,
    extract_file,
)

DATA_DIR = Path("data")
FILE_A1 = DATA_DIR / "corporates_A_1.xlsm"
FILE_A2 = DATA_DIR / "corporates_A_2.xlsm"
FILE_B1 = DATA_DIR / "corporates_B_1.xlsm"
FILE_B2 = DATA_DIR / "corporates_B_2.xlsm"


# ── helper: _build_label_map ──────────────────────────────────────────────────

class TestBuildLabelMap:
    def test_non_empty_labels_are_indexed(self):
        rows = [
            (None, "Label A", "value1"),
            (None, "Label B", "value2"),
        ]
        m = _build_label_map(rows)
        assert "Label A" in m
        assert "Label B" in m

    def test_none_labels_are_excluded(self):
        rows = [
            (None, None, "value"),
            (None, "Real label", "v"),
        ]
        m = _build_label_map(rows)
        assert None not in m
        assert "Real label" in m

    def test_empty_string_labels_are_excluded(self):
        rows = [(None, "   ", "value")]
        m = _build_label_map(rows)
        assert not m

    def test_label_is_stripped(self):
        rows = [(None, "  Label  ", "val")]
        m = _build_label_map(rows)
        assert "Label" in m

    def test_full_row_is_stored(self):
        row = (None, "Key", "v1", "v2", "v3")
        m = _build_label_map([row])
        assert m["Key"] == row


# ── helper: _extract_year_specs ───────────────────────────────────────────────

class TestExtractYearSpecs:
    def test_integer_years_are_historical(self):
        row = (None, "[Scope Credit Metrics]", 2020, 2021, 2022)
        specs = _extract_year_specs(row)
        assert specs == [(2020, False), (2021, False), (2022, False)]

    def test_estimate_strings_are_flagged(self):
        row = (None, "[Scope Credit Metrics]", 2023, "2024E", "2025E")
        specs = _extract_year_specs(row)
        assert specs == [(2023, False), (2024, True), (2025, True)]

    def test_stops_at_locked_sentinel(self):
        row = (None, "[Scope Credit Metrics]", 2022, 2023, "Locked", 2024)
        specs = _extract_year_specs(row)
        assert specs == [(2022, False), (2023, False)]

    def test_stops_at_none(self):
        row = (None, "[Scope Credit Metrics]", 2022, None, 2023)
        specs = _extract_year_specs(row)
        assert specs == [(2022, False)]

    def test_empty_row_returns_empty(self):
        row = (None, "[Scope Credit Metrics]")
        specs = _extract_year_specs(row)
        assert specs == []


# ── helper: _is_metric_row ────────────────────────────────────────────────────

class TestIsMetricRow:
    def test_numeric_col_c_is_metric(self):
        assert _is_metric_row((None, "some label", 3.14))

    def test_none_col_c_is_metric(self):
        assert _is_metric_row((None, "some label", None))

    def test_no_data_string_is_metric(self):
        assert _is_metric_row((None, "some label", "No data"))

    def test_locked_string_is_metric(self):
        assert _is_metric_row((None, "some label", "Locked"))

    def test_string_value_is_not_metric(self):
        assert not _is_metric_row((None, "Business risk profile", "BBB"))

    def test_short_row_is_not_metric(self):
        assert not _is_metric_row((None, "Label"))


# ── real-file tests ───────────────────────────────────────────────────────────

@pytest.mark.skipif(not FILE_A1.exists(), reason="data/corporates_A_1.xlsm not present")
class TestExtractFileA1:
    """Parse the real Company A v1 fixture and verify field values."""

    @pytest.fixture(scope="class")
    def result(self) -> ExtractedFile:
        return extract_file(FILE_A1)

    def test_returns_extracted_file(self, result):
        assert isinstance(result, ExtractedFile)

    def test_filename(self, result):
        assert result.filename == "corporates_A_1.xlsm"

    def test_file_hash_is_sha256_hex(self, result):
        assert len(result.file_hash) == 64
        assert all(c in "0123456789abcdef" for c in result.file_hash)

    def test_extracted_at_is_utc(self, result):
        assert result.extracted_at.tzinfo is not None

    def test_rated_entity(self, result):
        assert result.data.rated_entity == "Company A"

    def test_corporate_sector(self, result):
        assert result.data.corporate_sector == "Personal & Household Goods"

    def test_country_of_origin(self, result):
        assert result.data.country_of_origin == "Federal Republic of Germany"

    def test_reporting_currency(self, result):
        assert result.data.reporting_currency == "EUR"

    def test_accounting_principles(self, result):
        from src.core.constants import AccountingPrinciples
        assert result.data.accounting_principles == AccountingPrinciples.IFRS

    def test_business_year_end(self, result):
        from src.core.constants import BusinessYearEnd
        assert result.data.business_year_end == BusinessYearEnd.DECEMBER

    def test_rating_methodologies_non_empty(self, result):
        assert len(result.data.rating_methodologies) >= 1

    def test_single_industry_segment(self, result):
        assert len(result.data.industry_segments) == 1
        seg = result.data.industry_segments[0]
        assert seg.position == 1
        assert seg.industry_name == "Consumer Products: Non-Discretionary"
        assert seg.risk_score == "A"
        assert seg.weight == pytest.approx(1.0)

    def test_industry_weights_sum_to_one(self, result):
        total = sum(s.weight for s in result.data.industry_segments)
        assert total == pytest.approx(1.0, abs=0.01)

    def test_business_risk_profile(self, result):
        from src.core.constants import RatingGrade
        assert result.data.business_risk_profile == RatingGrade.B_PLUS

    def test_blended_industry_risk_profile_is_a(self, result):
        from src.core.constants import RatingGrade
        assert result.data.blended_industry_risk_profile == RatingGrade.A

    def test_scope_metrics_non_empty(self, result):
        assert len(result.data.scope_metrics) >= 1

    def test_scope_metrics_have_unique_name_year_estimate_combos(self, result):
        keys = [(m.metric_name, m.year, m.is_estimate) for m in result.data.scope_metrics]
        assert len(keys) == len(set(keys))

    def test_scope_metrics_estimates_flagged(self, result):
        estimates = [m for m in result.data.scope_metrics if m.is_estimate]
        assert len(estimates) >= 1

    def test_no_data_cells_become_none(self, result):
        # At least one cell has 'No data' in A1 fixtures
        none_metrics = [m for m in result.data.scope_metrics if m.value is None]
        assert len(none_metrics) >= 1


@pytest.mark.skipif(not FILE_B1.exists(), reason="data/corporates_B_1.xlsm not present")
class TestExtractFileB1:
    """Company B v1 has two industry segments — verify multi-segment handling."""

    @pytest.fixture(scope="class")
    def result(self) -> ExtractedFile:
        return extract_file(FILE_B1)

    def test_rated_entity(self, result):
        assert result.data.rated_entity == "Company B"

    def test_two_industry_segments(self, result):
        assert len(result.data.industry_segments) == 2

    def test_segment_positions(self, result):
        positions = [s.position for s in result.data.industry_segments]
        assert positions == [1, 2]

    def test_segment_weights_sum_to_one(self, result):
        total = sum(s.weight for s in result.data.industry_segments)
        assert total == pytest.approx(1.0, abs=0.01)

    def test_segment_names(self, result):
        names = [s.industry_name for s in result.data.industry_segments]
        assert "Automotive Suppliers" in names
        assert "Automotive and Commercial Vehicle Manufacturers" in names

    def test_single_methodology(self, result):
        assert len(result.data.rating_methodologies) == 1


@pytest.mark.skipif(not FILE_A2.exists(), reason="data/corporates_A_2.xlsm not present")
class TestExtractFileA2:
    """Company A v2 should differ from v1 in blended_industry_risk_profile."""

    @pytest.fixture(scope="class")
    def result(self) -> ExtractedFile:
        return extract_file(FILE_A2)

    def test_rated_entity_same_as_v1(self, result):
        assert result.data.rated_entity == "Company A"

    def test_industry_risk_score_is_bbb_in_v2(self, result):
        # A_1 industry risk score is 'A'; A_2 changes it to 'BBB'
        assert result.data.industry_segments[0].risk_score == "BBB"

    def test_single_methodology_in_v2(self, result):
        # A_1 has 2 methodologies; A_2 drops one
        assert len(result.data.rating_methodologies) == 1

    def test_hash_differs_from_a1(self, result):
        if not FILE_A1.exists():
            pytest.skip("A1 not present for comparison")
        a1 = extract_file(FILE_A1)
        assert result.file_hash != a1.file_hash


@pytest.mark.skipif(not FILE_B2.exists(), reason="data/corporates_B_2.xlsm not present")
class TestExtractFileB2:
    """Company B v2 changes segment weights from 0.15/0.85 → 0.25/0.75."""

    @pytest.fixture(scope="class")
    def result(self) -> ExtractedFile:
        return extract_file(FILE_B2)

    def test_rated_entity(self, result):
        assert result.data.rated_entity == "Company B"

    def test_two_industry_segments(self, result):
        assert len(result.data.industry_segments) == 2

    def test_segment_weights_changed_from_v1(self, result):
        weights = sorted(s.weight for s in result.data.industry_segments)
        assert weights == pytest.approx([0.25, 0.75], abs=0.001)

    def test_segment_weights_sum_to_one(self, result):
        total = sum(s.weight for s in result.data.industry_segments)
        assert total == pytest.approx(1.0, abs=0.01)

    def test_hash_differs_from_b1(self, result):
        if not FILE_B1.exists():
            pytest.skip("B1 not present for comparison")
        b1 = extract_file(FILE_B1)
        assert result.file_hash != b1.file_hash


# ── idempotency: same file → same hash ───────────────────────────────────────

@pytest.mark.skipif(not FILE_A1.exists(), reason="data/corporates_A_1.xlsm not present")
def test_extract_is_deterministic():
    """Extracting the same file twice yields the same hash."""
    r1 = extract_file(FILE_A1)
    r2 = extract_file(FILE_A1)
    assert r1.file_hash == r2.file_hash


# ── error-handling ────────────────────────────────────────────────────────────

class TestExtractionErrors:
    def test_missing_file_raises(self, tmp_path):
        with pytest.raises(ExtractionError, match="not found"):
            extract_file(tmp_path / "nonexistent.xlsm")

    def test_wrong_extension_raises(self, tmp_path):
        f = tmp_path / "data.xlsx"
        f.write_bytes(b"dummy")
        with pytest.raises(ExtractionError, match=".xlsm"):
            extract_file(f)

    def test_missing_master_sheet_raises(self, tmp_path):
        import openpyxl
        wb = openpyxl.Workbook()
        wb.active.title = "Sheet1"
        p = tmp_path / "test.xlsm"
        wb.save(p)
        with pytest.raises(ExtractionError, match="MASTER"):
            extract_file(p)

    def test_missing_required_label_raises(self, tmp_path):
        """A MASTER sheet that omits a required KV label raises ExtractionError."""
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "MASTER"
        # Provide only 'Rated entity'; every other required field is absent.
        ws.cell(row=2, column=2, value="Rated entity")
        ws.cell(row=2, column=3, value="Test Corp")
        p = tmp_path / "minimal.xlsm"
        wb.save(p)
        with pytest.raises(ExtractionError, match="Required field not found"):
            extract_file(p)

    def test_mismatched_industry_rows_raises(self, tmp_path):
        """Industry risk/score/weight rows with different column counts raise ExtractionError."""
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "MASTER"
        # Two industry names but only one score
        ws.cell(row=7, column=2, value="Industry risk")
        ws.cell(row=7, column=3, value="Segment A")
        ws.cell(row=7, column=4, value="Segment B")
        ws.cell(row=8, column=2, value="Industry risk score")
        ws.cell(row=8, column=3, value="BBB")
        ws.cell(row=9, column=2, value="Industry weight")
        ws.cell(row=9, column=3, value=0.5)
        ws.cell(row=9, column=4, value=0.5)
        p = tmp_path / "mismatch.xlsm"
        wb.save(p)
        with pytest.raises(ExtractionError, match="mismatched column counts"):
            from src.pipeline.extractor import _build_label_map, _parse_industry_segments
            rows = [
                tuple(ws.cell(row=r, column=c).value for c in range(1, 10))
                for r in range(1, 10)
            ]
            label_map = _build_label_map(rows)
            _parse_industry_segments(label_map)

    def test_no_metric_year_headers_raises(self, tmp_path):
        """A [Scope Credit Metrics] header with no year columns raises ExtractionError."""
        import openpyxl  # noqa: PLC0415

        from src.pipeline.extractor import _parse_scope_metrics
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "MASTER"
        ws.cell(row=1, column=2, value="[Scope Credit Metrics]")
        # No year values after the header
        p = tmp_path / "noyears.xlsm"
        wb.save(p)
        rows = [(None, "[Scope Credit Metrics]")]
        with pytest.raises(ExtractionError, match="No year columns"):
            _parse_scope_metrics(rows)

    def test_non_numeric_metric_value_raises_value_error(self):
        """A metric row where col C is numeric but a later column is garbage raises ValueError.

        _is_metric_row passes (col C is float), but float("GARBAGE") fails for col D.
        extract_file wraps this ValueError as ExtractionError (fix 2).
        """
        from src.pipeline.extractor import _parse_scope_metrics
        rows = [
            (None, "[Scope Credit Metrics]", 2022, 2023),
            # col C = 3.14 passes _is_metric_row; col D = "GARBAGE" fails float()
            (None, "My Metric", 3.14, "GARBAGE"),
        ]
        with pytest.raises(ValueError):
            _parse_scope_metrics(rows)
