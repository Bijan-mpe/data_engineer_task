"""
Extractor: reads one .xlsm source file and returns an ExtractedFile DTO.

Single responsibility: open the workbook, parse the MASTER sheet key-value
structure, compute the SHA-256 file hash, and assemble a fully-populated
ExtractedFile.  No validation, transformation, or database concerns here.

MASTER sheet layout
-------------------
The sheet is a non-tabular key-value document.  Column B (index 1) holds
field labels; column C onwards (index 2+) holds values.  Some fields span
multiple columns for multi-segment or multi-methodology companies:

  Row  2  Rated entity           → single value
  Row  5  Rating methodologies   → one per column (1..N)
  Rows 7-9  Industry risk/score/weight → one column per segment
  Rows 12-15  Reporting currency, country, accounting, year-end
  Rows 18-31  Business & financial risk rating grades
  Row 35  [Scope Credit Metrics] header, year labels in cols C onwards
  Rows 36-41  One metric per row, values aligned with header columns

Parsing is label-driven (MasterField enum) rather than row-number-driven,
which insulates the code from minor row insertions in future file versions.
"""

from __future__ import annotations

import hashlib
import re
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import openpyxl
from pydantic import ValidationError

from src.core.constants import DATA_FILE_EXTENSION, MASTER_SHEET_NAME, MasterField
from src.core.logging import get_logger
from src.models.schemas import (
    ExtractedFile,
    IndustrySegment,
    RawMasterData,
    ScopeMetric,
)

logger = get_logger(__name__)

# Marker string at the end of each metric row's column range (excluded from values).
_LOCKED_SENTINEL = "Locked"
_NO_DATA_SENTINEL = "No data"

# Regex matching forward-estimate year labels like "2025E", "2026E".
_ESTIMATE_RE = re.compile(r"^(\d{4})E$")


class ExtractionError(Exception):
    """Raised when the MASTER sheet cannot be parsed as expected.

    Wraps the source path and a human-readable reason so the pipeline can
    log a clear message without needing to catch generic exceptions.
    """


def extract_file(path: Path) -> ExtractedFile:
    """Parse one .xlsm file and return its content as an ExtractedFile DTO.

    Steps:
    1. Verify the file exists and has the expected extension.
    2. Compute the SHA-256 hash of the raw bytes.
    3. Open the workbook (read-only, VBA preserved to avoid openpyxl warnings).
    4. Locate the MASTER sheet and build a label→row mapping.
    5. Parse each field group from the mapping.
    6. Return an ExtractedFile wrapping the assembled RawMasterData.

    Parameters
    ----------
    path:
        Absolute or relative path to a .xlsm source file.

    Returns
    -------
    ExtractedFile
        Fully populated DTO ready for the validator stage.

    Raises
    ------
    ExtractionError
        If the file is missing, has the wrong extension, lacks the MASTER
        sheet, or is missing any required field label.
    """
    path = Path(path)
    if not path.exists():
        raise ExtractionError(f"File not found: {path}")
    if path.suffix.lower() != DATA_FILE_EXTENSION:
        raise ExtractionError(
            f"Expected {DATA_FILE_EXTENSION} file, got: {path.suffix}"
        )

    logger.bind(filename=path.name).info("extractor.file_started")
    file_hash = _compute_sha256(path)
    try:
        raw_data = _parse_master_sheet(path)
    except ExtractionError:
        raise
    except (ValidationError, ValueError) as exc:
        raise ExtractionError(
            f"[{path.name}] extracted values failed schema validation: {exc}"
        ) from exc

    return ExtractedFile(
        filename=path.name,
        file_hash=file_hash,
        extracted_at=datetime.now(tz=timezone.utc),
        data=raw_data,
    )


# ── private helpers ───────────────────────────────────────────────────────────


def _compute_sha256(path: Path) -> str:
    """Return the lowercase hex SHA-256 digest of the file at *path*."""
    h = hashlib.sha256()
    with path.open("rb") as fh:
        for chunk in iter(lambda: fh.read(65536), b""):
            h.update(chunk)
    return h.hexdigest()


def _parse_master_sheet(path: Path) -> RawMasterData:
    """Open *path*, read the MASTER sheet, and return a RawMasterData DTO.

    Parameters
    ----------
    path:
        Path to the .xlsm workbook.

    Returns
    -------
    RawMasterData
        Populated from the MASTER sheet KV structure.

    Raises
    ------
    ExtractionError
        If the MASTER sheet is absent or a required label is not found.
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True, keep_vba=True)
    try:
        if MASTER_SHEET_NAME not in wb.sheetnames:
            raise ExtractionError(
                f"Workbook '{path.name}' has no '{MASTER_SHEET_NAME}' sheet. "
                f"Found: {wb.sheetnames}"
            )
        ws = wb[MASTER_SHEET_NAME]
        rows: list[tuple[Any, ...]] = [
            tuple(cell for cell in row) for row in ws.iter_rows(values_only=True)
        ]
    finally:
        wb.close()

    label_map = _build_label_map(rows)
    return _assemble_raw_data(label_map, rows, path.name)


def _build_label_map(rows: list[tuple[Any, ...]]) -> dict[str, tuple[Any, ...]]:
    """Return a mapping from label (col B) to the full row tuple.

    Parsing stops when the '[Scope Credit Metrics]' header is encountered so
    that the 'Liquidity' scope-metric row (which shares a label with the KV
    Liquidity field) never overwrites the KV entry.  First occurrence wins,
    so duplicate labels earlier in the sheet are handled safely too.
    """
    label_map: dict[str, tuple[Any, ...]] = {}
    for row in rows:
        if len(row) < 2:
            continue
        label = row[1]
        if not isinstance(label, str) or not label.strip():
            continue
        stripped = label.strip()
        if stripped == MasterField.SCOPE_CREDIT_METRICS.value:
            break  # everything from here on belongs to the metrics section
        if stripped not in label_map:  # first-wins: ignore duplicate labels
            label_map[stripped] = row
    return label_map


def _require(label_map: dict[str, tuple[Any, ...]], field: MasterField) -> tuple[Any, ...]:
    """Return the row for *field* or raise ExtractionError if absent."""
    row = label_map.get(field.value)
    if row is None:
        raise ExtractionError(f"Required field not found in MASTER sheet: '{field.value}'")
    return row


def _single_value(
    label_map: dict[str, tuple[Any, ...]], field: MasterField
) -> Any:
    """Return the primary (column-C) value for *field*."""
    row = _require(label_map, field)
    return row[2] if len(row) > 2 else None


def _multi_values(
    label_map: dict[str, tuple[Any, ...]], field: MasterField
) -> list[Any]:
    """Return all non-None values starting at column C for *field*."""
    row = _require(label_map, field)
    return [v for v in row[2:] if v is not None]


def _assemble_raw_data(
    label_map: dict[str, tuple[Any, ...]],
    rows: list[tuple[Any, ...]],
    filename: str,
) -> RawMasterData:
    """Build a RawMasterData from the KV label_map and raw rows.

    Parameters
    ----------
    label_map:
        KV mapping produced by _build_label_map (stops before scope metrics).
    rows:
        Full list of rows from the MASTER sheet, used for scope metrics parsing.
    filename:
        Used only in ExtractionError messages for context.
    """
    # ── industry segments (rows 7-9) ─────────────────────────────────────────
    industry_segments = _parse_industry_segments(label_map)

    # ── rating methodologies (row 5) ─────────────────────────────────────────
    rating_methodologies = _multi_values(label_map, MasterField.RATING_METHODOLOGIES)
    rating_methodologies = [str(m) for m in rating_methodologies if m]

    # ── single scalar fields ──────────────────────────────────────────────────
    def _str(field: MasterField) -> str:
        v = _single_value(label_map, field)
        if v is None:
            raise ExtractionError(
                f"[{filename}] field '{field.value}' has no value in column C"
            )
        return str(v).strip()

    def _optional_str(field: MasterField) -> str | None:
        v = _single_value(label_map, field)
        if v is None or str(v).strip() == "":
            return None
        return str(v).strip()

    # ── scope metrics (rows 35+) ──────────────────────────────────────────────
    scope_metrics = _parse_scope_metrics(rows)

    return RawMasterData(
        rated_entity=_str(MasterField.RATED_ENTITY),
        corporate_sector=_str(MasterField.CORPORATE_SECTOR),
        rating_methodologies=rating_methodologies,
        industry_segments=industry_segments,
        segmentation_criteria=_optional_str(MasterField.SEGMENTATION_CRITERIA),
        reporting_currency=_str(MasterField.REPORTING_CURRENCY),
        country_of_origin=_str(MasterField.COUNTRY_OF_ORIGIN),
        accounting_principles=_str(MasterField.ACCOUNTING_PRINCIPLES),
        business_year_end=_str(MasterField.BUSINESS_YEAR_END),
        business_risk_profile=_str(MasterField.BUSINESS_RISK_PROFILE),
        blended_industry_risk_profile=_str(MasterField.BLENDED_INDUSTRY_RISK_PROFILE),
        competitive_positioning=_str(MasterField.COMPETITIVE_POSITIONING),
        market_share=_str(MasterField.MARKET_SHARE),
        diversification=_str(MasterField.DIVERSIFICATION),
        operating_profitability=_str(MasterField.OPERATING_PROFITABILITY),
        sector_specific_factor_1=_optional_str(MasterField.SECTOR_SPECIFIC_1),
        sector_specific_factor_2=_optional_str(MasterField.SECTOR_SPECIFIC_2),
        financial_risk_profile=_str(MasterField.FINANCIAL_RISK_PROFILE),
        leverage=_str(MasterField.LEVERAGE),
        interest_cover=_str(MasterField.INTEREST_COVER),
        cash_flow_cover=_str(MasterField.CASH_FLOW_COVER),
        liquidity=_str(MasterField.LIQUIDITY),
        scope_metrics=scope_metrics,
    )


def _parse_industry_segments(
    label_map: dict[str, tuple[Any, ...]]
) -> list[IndustrySegment]:
    """Extract industry segments from the three parallel MASTER rows.

    Rows 7, 8, and 9 each hold one value per segment column (C, D, E, …).
    The number of segments equals the count of non-None values in the
    'Industry risk' row.
    """
    names = _multi_values(label_map, MasterField.INDUSTRY_RISK)
    scores = _multi_values(label_map, MasterField.INDUSTRY_RISK_SCORE)
    weights = _multi_values(label_map, MasterField.INDUSTRY_WEIGHT)

    if not (len(names) == len(scores) == len(weights)):
        raise ExtractionError(
            f"Industry segment rows have mismatched column counts: "
            f"names={len(names)}, scores={len(scores)}, weights={len(weights)}"
        )

    return [
        IndustrySegment(
            position=i + 1,
            industry_name=str(names[i]).strip(),
            risk_score=str(scores[i]).strip(),
            weight=float(weights[i]),
        )
        for i in range(len(names))
    ]


def _parse_scope_metrics(rows: list[tuple[Any, ...]]) -> list[ScopeMetric]:
    """Parse the [Scope Credit Metrics] section from the raw row list.

    Scans rows for the '[Scope Credit Metrics]' header, reads year labels
    from that row (cols C onwards), then collects all subsequent non-empty
    rows as metric observations.

    Using raw rows (not the KV label_map) avoids the label collision with
    the 'Liquidity' KV field: both share the same column-B text but appear
    in different sections of the sheet.
    """
    header_idx: int | None = None
    for i, row in enumerate(rows):
        if len(row) >= 2 and row[1] == MasterField.SCOPE_CREDIT_METRICS.value:
            header_idx = i
            break

    if header_idx is None:
        raise ExtractionError(
            f"Required field not found in MASTER sheet: '{MasterField.SCOPE_CREDIT_METRICS.value}'"
        )

    year_specs = _extract_year_specs(rows[header_idx])
    if not year_specs:
        raise ExtractionError("No year columns found in [Scope Credit Metrics] header row")

    metrics: list[ScopeMetric] = []
    for row in rows[header_idx + 1:]:
        if len(row) < 2:
            continue
        label = row[1]
        if not isinstance(label, str) or not label.strip():
            continue
        if not _is_metric_row(row):
            continue  # skip footer/note rows whose col-C value is a plain string
        metric_name = label.strip()
        for col_offset, (year, is_estimate) in enumerate(year_specs):
            raw = row[2 + col_offset] if (2 + col_offset) < len(row) else None
            if raw == _LOCKED_SENTINEL:
                break
            value: float | None
            if raw is None or raw == _NO_DATA_SENTINEL:
                value = None
            else:
                value = float(raw)
            metrics.append(
                ScopeMetric(
                    metric_name=metric_name,
                    year=year,
                    is_estimate=is_estimate,
                    value=value,
                )
            )

    return metrics


def _extract_year_specs(header_row: tuple[Any, ...]) -> list[tuple[int, bool]]:
    """Parse year labels from the header row into (year, is_estimate) pairs.

    Stops at the first 'Locked' sentinel or None value after a valid year.
    """
    year_specs: list[tuple[int, bool]] = []
    for cell in header_row[2:]:
        if cell is None or cell == _LOCKED_SENTINEL:
            break
        if isinstance(cell, int):
            year_specs.append((cell, False))
        elif isinstance(cell, str):
            m = _ESTIMATE_RE.match(cell.strip())
            if m:
                year_specs.append((int(m.group(1)), True))
    return year_specs


def _is_metric_row(row: tuple[Any, ...]) -> bool:
    """Return True if the row's first data column looks like a metric value.

    A metric row has a numeric value, None, 'No data', or 'Locked' at col C
    (index 2) — not a string that would be another KV label.
    """
    if len(row) < 3:
        return False
    val = row[2]
    if val is None:
        return True
    if isinstance(val, (int, float)):
        return True
    if isinstance(val, str) and val in (_NO_DATA_SENTINEL, _LOCKED_SENTINEL):
        return True
    return False
